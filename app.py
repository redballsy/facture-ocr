import sys
import os
import re
import io
import base64
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

from flask import Flask, render_template, request, jsonify, send_file

try:
    import pytesseract
    from PIL import Image, ImageEnhance, ImageFilter, ImageDraw, ImageFont
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from pdf2image import convert_from_bytes
except ImportError as e:
    print(f"Dependance manquante: {e}")
    sys.exit(1)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "gesfi-secret-key-2026")

EXTENSIONS = {"png", "pdf"}
COULEURS_GESFI = {"primary": "#1B3A5C", "secondary": "#2E5A88", "accent": "#D4AF37"}

CHEMINS_TESSERACT = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    "/usr/bin/tesseract",
    "/usr/local/bin/tesseract",
]

def configurer_tesseract():
    for chemin in CHEMINS_TESSERACT:
        if Path(chemin).exists():
            pytesseract.pytesseract.tesseract_cmd = chemin
            return True
    try:
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False

TESSERACT_OK = configurer_tesseract()

def extension_autorisee(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in EXTENSIONS

# ── OCR ───────────────────────────────────────────────────────────────────────

def pretraiter_image(img_bytes: bytes) -> Image.Image:
    img = Image.open(io.BytesIO(img_bytes))
    if img.mode != 'RGB':
        img = img.convert('RGB')
    if img.width < 1400:
        ratio = 1400 / img.width
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
    img = img.convert('L')
    img = ImageEnhance.Contrast(img).enhance(2.5)
    img = img.filter(ImageFilter.SHARPEN)
    return img

def extraire_texte_image(img_bytes: bytes) -> str:
    img = pretraiter_image(img_bytes)
    for config in ["--oem 3 --psm 6", "--oem 3 --psm 4"]:
        try:
            t = pytesseract.image_to_string(img, lang="fra+eng", config=config)
            if t.strip():
                return t
        except Exception:
            pass
    return ""

def traiter_pdf(img_bytes: bytes) -> str:
    try:
        images = convert_from_bytes(img_bytes, dpi=300)
        out = ""
        for i, image in enumerate(images):
            buf = io.BytesIO()
            image.save(buf, format='PNG')
            out += f"\n--- Page {i+1} ---\n{extraire_texte_image(buf.getvalue())}\n"
        return out
    except Exception:
        return ""

def extraire_texte(img_bytes: bytes, filename: str) -> str:
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    return traiter_pdf(img_bytes) if ext == 'pdf' else extraire_texte_image(img_bytes)

# ── Utilitaires ───────────────────────────────────────────────────────────────

def net(s):
    return s.strip() if s else ""

def chercher(pattern, texte, groupe=1, defaut=""):
    m = re.search(pattern, texte, re.IGNORECASE | re.MULTILINE | re.DOTALL)
    return net(m.group(groupe)) if m else defaut

def to_float(s: str) -> float:
    """Convertit un montant texte (avec espaces, virgules, points) en float."""
    if not s:
        return 0.0
    s = re.sub(r'[^\d,\.]', '', s.replace('\u202f', '').replace('\xa0', '').replace(' ', ''))
    # Format français : 200.000,00 (point = milliers, virgule = décimale)
    if re.match(r'^\d{1,3}(\.\d{3})+(,\d+)?$', s):
        s = s.replace('.', '').replace(',', '.')
    # Virgule seule = décimale
    elif ',' in s and '.' not in s:
        s = s.replace(',', '.')
    # Point avec 3 décimales = séparateur milliers : 200.000 → 200000
    elif '.' in s and len(s.split('.')[-1]) == 3:
        s = s.replace('.', '')
    try:
        return float(s)
    except ValueError:
        return 0.0

def fmt_fcfa(val: float) -> str:
    return f"{val:,.0f} FCFA".replace(",", " ")

def calculer_totaux(prix_ht: float, quantite: float, tva_pct: float) -> dict:
    sous_total_ht = prix_ht * quantite
    montant_tva   = sous_total_ht * (tva_pct / 100) if tva_pct > 0 else 0.0
    total_ttc     = sous_total_ht + montant_tva
    return {
        "prix_ht":      prix_ht,
        "quantite":     quantite,
        "sous_total_ht": sous_total_ht,
        "tva_pct":      tva_pct,
        "montant_tva":  montant_tva,
        "total_ttc":    total_ttc,
    }

def detecter_type_facture(texte: str) -> str:
    tl = texte.lower()
    if "facture d'achat" in tl or "fournisseur" in tl:
        return "achat"
    return "vente"

# ── Extraction robuste du bloc fournisseur ────────────────────────────────────

def extraire_bloc_fournisseur(texte: str) -> dict:
    """
    Extrait nom, adresse, tel, email du bloc FOURNISSEUR.
    S'arrête avant la première ligne qui ressemble à un article ou à un total.
    """
    result = {"fournisseur_nom": "", "fournisseur_adresse": "", 
              "fournisseur_tel": "", "fournisseur_email": ""}

    # Extraire le bloc entre FOURNISSEUR et la prochaine section (article/total/Description)
    fblock_m = re.search(
        r'FOURNISSEUR\s*\n(.*?)(?=\nDescription|\nImprimante|\nTablette|\nOrdinateur'
        r'|\nSmartphone|\nSous-total|\nTVA\s*:|\nTotal|\n[A-Z]{2,}\s*\n|$)',
        texte, re.S | re.I
    )
    if not fblock_m:
        return result

    bloc = fblock_m.group(1)
    lignes = [l.strip() for l in bloc.splitlines() if l.strip()]

    for ligne in lignes:
        if re.match(r'T[eé]l\s*:', ligne, re.I):
            result["fournisseur_tel"] = re.sub(r'T[eé]l\s*:\s*', '', ligne, flags=re.I).strip()
        elif re.match(r'Email\s*:', ligne, re.I):
            result["fournisseur_email"] = re.sub(r'Email\s*:\s*', '', ligne, flags=re.I).strip()
        elif not result["fournisseur_nom"]:
            result["fournisseur_nom"] = ligne
        elif not result["fournisseur_adresse"]:
            result["fournisseur_adresse"] = ligne

    return result

# ── Extraction robuste de la source ──────────────────────────────────────────

def extraire_source(texte: str) -> str:
    """
    Cherche l'adresse source (ex: "08 BP ABIDJAN...").
    Priorité : ligne contenant "BP" puis regex générique après "Source".
    """
    # Ligne avec BP ABIDJAN (adresse postale ivoirienne typique)
    bp_m = re.search(r'(\d+\s*BP\s+[A-Z][^\n\r]+)', texte, re.I)
    if bp_m:
        # Couper avant "Livraison" ou "Date" si sur la même ligne
        src = re.split(r'\s+(?:Livraison|Date)\s', bp_m.group(1))[0]
        return src.strip()

    # Fallback : ligne après "Source"
    src_m = re.search(r'Source\s*\n\s*(.+)', texte, re.I)
    if src_m:
        return src_m.group(1).strip()
    return ""

# ── Extraction robuste du tableau article ────────────────────────────────────

def extraire_tableau_article(texte: str) -> dict:
    """
    Extrait description, quantite, prix_ht, tva_pct.

    Stratégie :
    1. Trouver l'en-tête du tableau (ligne contenant "Description" + "Qté"/"Prix")
    2. Lire la(les) ligne(s) article après l'en-tête
    3. Prix = total_ligne / quantité (robuste face aux corruptions OCR du prix unitaire)
    4. Fallback sur Sous-total HT si rien ne marche
    """
    lignes = texte.splitlines()
    result = {"description": "", "quantite": 1.0, "prix_ht": 0.0, "tva_pct": 0.0}

    # ── Trouver l'en-tête ────────────────────────────────────────────────────
    header_idx = -1
    for i, ligne in enumerate(lignes):
        if re.search(r'description', ligne, re.I) and \
           re.search(r'(qt[eé]|qte|quantit|prix|total)', ligne, re.I):
            header_idx = i
            break

    # ── Lignes article ───────────────────────────────────────────────────────
    article_lignes = []
    if header_idx >= 0:
        for ligne in lignes[header_idx + 1:]:
            s = ligne.strip()
            if not s:
                continue
            if re.search(r'(sous.total|total\s+ttc|total\s+ht|^tva\s*:)', s, re.I):
                break
            article_lignes.append(s)

    # ── Parser chaque ligne article ──────────────────────────────────────────
    for ligne in article_lignes:
        # Description : texte en début de ligne
        desc_m = re.match(r'^([A-Za-z\u00C0-\u00FF][A-Za-z\u00C0-\u00FF\s\-\'\u2019\.]+)', ligne)
        if not desc_m or len(desc_m.group(1).strip()) < 3:
            continue
        desc = desc_m.group(1).strip()

        # Quantité : premier nombre juste après la description
        reste = ligne[desc_m.end():]
        qte_m = re.match(r'\s*(\d+[,\.]?\d*)', reste)
        qte = to_float(qte_m.group(1)) if qte_m else 1.0
        if qte <= 0:
            qte = 1.0

        # TVA % sur la ligne
        tva_m = re.search(r'(\d+)\s*%', ligne)
        tva_pct = float(tva_m.group(1)) if tva_m else 0.0

        # Total ligne = dernier montant FCFA sur la ligne
        # (plus fiable que le prix unitaire qui peut être corrompu par OCR)
        montants_fcfa = re.findall(
            r'(\d[\d\s]*[\d](?:[,\.]\d{2})?(?=\s*FCFA|\s*$))', ligne
        )
        total_ligne_val = 0.0
        for m in reversed(montants_fcfa):
            v = to_float(m)
            if v > 100:
                total_ligne_val = v
                break

        # Prix unitaire HT = total_ligne / quantite
        if total_ligne_val > 0:
            prix_ht = total_ligne_val / qte
        else:
            # Fallback : chercher un montant > 100 ailleurs dans la ligne
            apres_qte = reste[qte_m.end():] if qte_m else reste
            montants2 = re.findall(r'(\d[\d\s]*[\d][,\.]\d{2})', apres_qte)
            prix_candidates = [to_float(m) for m in montants2 if to_float(m) > 100]
            prix_ht = prix_candidates[0] if prix_candidates else 0.0

        result["description"] = desc
        result["quantite"]    = qte
        result["prix_ht"]     = prix_ht
        result["tva_pct"]     = tva_pct
        break  # première ligne article valide suffit

    # ── Fallback : Sous-total HT de la facture ───────────────────────────────
    if result["prix_ht"] == 0.0:
        sous_m = re.search(
            r'Sous.total\s+HT\s*:?\s*\**\s*([\d\s\.,]+)\s*FCFA', texte, re.I
        )
        if sous_m:
            st = to_float(sous_m.group(1))
            if st > 0:
                result["prix_ht"] = st / max(result["quantite"], 1.0)

    if result["prix_ht"] == 0.0:
        ttc_m = re.search(
            r'Total\s+TTC\s*:?\s*\**\s*([\d\s\.,]+)\s*FCFA', texte, re.I
        )
        if ttc_m:
            result["prix_ht"] = to_float(ttc_m.group(1)) / max(result["quantite"], 1.0)

    # ── TVA % globale fallback ────────────────────────────────────────────────
    if result["tva_pct"] == 0.0:
        tva_col = re.search(r'TVA\s+(\d+)\s*%', texte, re.I)
        if tva_col:
            result["tva_pct"] = float(tva_col.group(1))

    return result

# ── Extraction facture ACHAT ──────────────────────────────────────────────────

def extraire_facture_achat(texte: str) -> dict:
    d = {}

    # Numéro : chercher PO-XXXX directement
    po_m = re.search(r'(PO-\d+)', texte)
    if po_m:
        d["num_facture"] = po_m.group(1)
    else:
        num = chercher(r'N[°\u00b0]?\s*([\w\-]+)', texte)
        d["num_facture"] = num or ""

    # Dates
    d["date"]      = chercher(r'Date\s*:\s*([^\n\r]+)', texte)
    d["livraison"] = chercher(r'Livraison\s+pr[eé]vue\s*:\s*([^\n\r]+)', texte)

    # Source
    d["source"] = extraire_source(texte)

    # Fournisseur
    d.update(extraire_bloc_fournisseur(texte))

    # Article
    art = extraire_tableau_article(texte)

    # Si description encore vide, chercher les articles connus directement
    if not art["description"]:
        # Chercher un article reconnaissable sur une seule ligne
        for ligne in texte.splitlines():
            m = re.search(
                r'^.*?(Imprimante\s+de\s+bureau|Ordinateur|Tablette|Smartphone'
                r'|T[eé]l[eé]phone|Clavier|Souris|Ecran|Moniteur)'
                r'\s+([\d,\.]+).*?(\d[\d\s\.,]+)\s*FCFA',
                ligne, re.I
            )
            if m:
                art["description"] = m.group(1).strip()
                art["quantite"]    = to_float(m.group(2))
                total_l            = to_float(m.group(3))
                art["prix_ht"]     = total_l / max(art["quantite"], 1.0)
                break
        # Fallback générique : première ligne avec un grand montant FCFA
        if not art["description"]:
            for ligne in texte.splitlines():
                m = re.match(r'^([A-Z][a-zA-Z\u00C0-\u00FF][a-zA-Z\u00C0-\u00FF\s]{3,30})\s+([\d,\.]+).*?(\d[\d\s\.,]{3,})\s*FCFA', ligne)
                if m and to_float(m.group(3)) > 1000:
                    art["description"] = m.group(1).strip()
                    art["quantite"]    = to_float(m.group(2))
                    art["prix_ht"]     = to_float(m.group(3)) / max(art["quantite"], 1.0)
                    break

    d.update(art)
    totaux = calculer_totaux(d["prix_ht"], d["quantite"], d["tva_pct"])
    d.update(totaux)
    d["quantite_str"] = f"{d['quantite']:.2f}".rstrip('0').rstrip('.')
    return d

# ── Extraction facture VENTE ──────────────────────────────────────────────────

def extraire_facture_vente(texte: str) -> dict:
    d = {}

    # Numéro
    inv_m = re.search(r'(INV-[\d]+)', texte, re.I)
    d["num_facture"] = inv_m.group(1) if inv_m else chercher(r'FACTURE\s+([A-Z0-9\-]+)', texte)

    # Dates
    d["date_emission"] = chercher(r"Date\s+d['\u2019\u0060]?[\u00e9e]mission\s*:\s*([^\n\r]+)", texte)
    d["date_echeance"] = chercher(r"Date\s+d['\u2019\u0060]?[\u00e9e]ch[e\u00e9]ance\s*:\s*([^\n\r]+)", texte)

    # Source
    d["source"] = extraire_source(texte)

    # RCCM / Capital
    d["rccm"]    = chercher(r'(RCCM\s+CI[-\s]+[A-Z0-9\-\s]+)', texte).strip()
    cap = chercher(r'CAPITAL\s+([\d\s]+)', texte)
    d["capital"] = cap.replace(' ', '') if cap else ""

    # Client
    d["client"]   = chercher(r'CLIENT\s*\n\s*([^\n]+)', texte)
    d["paiement"] = chercher(r'PAIEMENT\s*(?:WAVE)?\s*[:\-]?\s*([^\n\r]+)', texte)

    # Article
    art = extraire_tableau_article(texte)
    d.update(art)
    totaux = calculer_totaux(d["prix_ht"], d["quantite"], d["tva_pct"])
    d.update(totaux)
    d["quantite_str"] = f"{d['quantite']:.2f}".rstrip('0').rstrip('.')
    return d

# ── Parser principal ──────────────────────────────────────────────────────────

def parser_facture_complet(texte: str, nom_fichier: str, solde: bool = False) -> dict:
    type_facture = detecter_type_facture(texte)
    if type_facture == "achat":
        donnees = extraire_facture_achat(texte)
        donnees["type"] = "ACHAT"
    else:
        donnees = extraire_facture_vente(texte)
        donnees["type"] = "VENTE"
    donnees["nom_fichier"] = nom_fichier
    donnees["solde"] = solde
    return donnees

# ── Génération image de synthèse ─────────────────────────────────────────────

def generer_image_facture(donnees: dict) -> str:
    W = 860

    C_DARK   = (27, 58, 92)
    C_ACCENT = (212, 175, 55)
    C_WHITE  = (255, 255, 255)
    C_LIGHT  = (232, 239, 248)
    C_GREEN  = (30, 130, 60)
    C_RED    = (170, 30, 30)
    C_GRAY   = (95, 108, 126)
    C_TEXT   = (25, 38, 58)
    C_SEC    = (42, 82, 130)
    C_BG     = (244, 246, 251)

    def load_font(bold=False, size=13):
        try:
            name = "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"
            return ImageFont.truetype(f"/usr/share/fonts/truetype/dejavu/{name}", size)
        except Exception:
            return ImageFont.load_default()

    f_title = load_font(True,  22)
    f_bold  = load_font(True,  14)
    f_reg   = load_font(False, 13)
    f_small = load_font(False, 11)
    f_large = load_font(True,  20)

    type_label = donnees.get("type", "VENTE")

    # Préparer les lignes pour calculer la hauteur
    rows_gen = []
    if type_label == "ACHAT":
        for lbl, key in [("Date", "date"), ("Livraison prevue", "livraison"), ("Source", "source")]:
            if donnees.get(key):
                rows_gen.append((lbl, donnees[key]))
        rows_inter = []
        for lbl, key in [("Nom", "fournisseur_nom"), ("Adresse", "fournisseur_adresse"),
                         ("Telephone", "fournisseur_tel"), ("Email", "fournisseur_email")]:
            if donnees.get(key):
                rows_inter.append((lbl, donnees[key]))
        sec_inter = "  FOURNISSEUR"
    else:
        for lbl, key in [("Date d emission", "date_emission"), ("Date d echeance", "date_echeance"),
                         ("Source", "source"), ("RCCM", "rccm"), ("Capital", "capital")]:
            if donnees.get(key):
                rows_gen.append((lbl, donnees[key]))
        rows_inter = []
        if donnees.get("client"):
            rows_inter.append(("Client", donnees["client"]))
        if donnees.get("paiement"):
            rows_inter.append(("Paiement", donnees["paiement"]))
        sec_inter = "  CLIENT"

    tva_pct = donnees.get("tva_pct", 0.0)
    n_fin   = 3 if tva_pct > 0 else 2

    ROW_H  = 32
    SEC_H  = 34
    SEP_H  = 5
    HEADER = 74
    FOOTER = 34

    H = (HEADER + 8
         + SEC_H + len(rows_gen)   * ROW_H + SEP_H
         + SEC_H + len(rows_inter) * ROW_H + SEP_H
         + SEC_H + 3 * ROW_H + SEP_H
         + SEC_H + n_fin * ROW_H + 50
         + FOOTER + 8)

    img  = Image.new("RGB", (W, H), color=C_BG)
    draw = ImageDraw.Draw(img)

    y = 0

    # En-tête
    draw.rectangle([(0, 0), (W, HEADER)], fill=C_DARK)
    draw.text((18, 10), "GESFI GROUP", font=f_title, fill=C_ACCENT)
    num = donnees.get("num_facture", "—")
    draw.text((18, 42), f"FACTURE D'{type_label}   \u2022   {num}", font=f_reg, fill=C_WHITE)

    solde    = donnees.get("solde", False)
    badge_c  = C_GREEN if solde else C_RED
    badge_t  = " SOLDE " if solde else " NON SOLDE "
    try:
        bb = draw.textbbox((0,0), badge_t, font=f_bold)
        bw = bb[2]-bb[0] + 20
    except Exception:
        bw = len(badge_t)*8 + 20
    bx = W - bw - 14
    draw.rectangle([(bx, 16), (W-14, 56)], fill=badge_c)
    draw.text((bx + 10, 28), badge_t, font=f_bold, fill=C_WHITE)

    y = HEADER + 8

    def section(title, color=C_DARK):
        nonlocal y
        draw.rectangle([(0, y), (W, y+SEC_H)], fill=color)
        draw.text((14, y+(SEC_H-14)//2), title, font=f_bold, fill=C_WHITE)
        y += SEC_H

    def row(label, value, alt=False):
        nonlocal y
        val_str = str(value).strip()
        if not val_str:
            return
        draw.rectangle([(0, y), (W, y+ROW_H)], fill=(C_LIGHT if alt else C_WHITE))
        draw.text((14,  y+(ROW_H-13)//2), str(label),        font=f_bold, fill=C_GRAY)
        draw.text((290, y+(ROW_H-13)//2), val_str[:72],      font=f_reg,  fill=C_TEXT)
        y += ROW_H

    def sep():
        nonlocal y
        draw.rectangle([(0, y), (W, y+SEP_H)], fill=C_ACCENT)
        y += SEP_H

    # Infos générales
    section("  INFORMATIONS GENERALES")
    for i, (lbl, val) in enumerate(rows_gen):
        row(lbl, val, i%2==1)
    sep()

    # Interlocuteur
    section(sec_inter, C_SEC)
    for i, (lbl, val) in enumerate(rows_inter):
        row(lbl, val, i%2==1)
    sep()

    # Détail article
    section("  DETAIL ARTICLE", C_SEC)
    row("Description",       donnees.get("description", "—"),           False)
    row("Quantite",          donnees.get("quantite_str", "1"),           True)
    row("Prix unitaire HT",  fmt_fcfa(donnees.get("prix_ht", 0.0)),     False)
    sep()

    # Calcul financier
    section("  CALCUL FINANCIER")
    sous_total  = donnees.get("sous_total_ht", 0.0)
    montant_tva = donnees.get("montant_tva",   0.0)
    total_ttc   = donnees.get("total_ttc",     0.0)
    qte_s       = donnees.get("quantite_str",  "1")
    prix_u      = donnees.get("prix_ht",       0.0)

    row(f"Sous-total HT  ({qte_s} x {fmt_fcfa(prix_u)})", fmt_fcfa(sous_total), False)
    if tva_pct > 0:
        row(f"TVA ({tva_pct:.0f}%)  =  {fmt_fcfa(sous_total)} x {tva_pct:.0f}%",
            fmt_fcfa(montant_tva), True)
    else:
        row("TVA", "0% — exonere", True)

    # Total TTC
    draw.rectangle([(0, y), (W, y+50)], fill=C_ACCENT)
    draw.text((16, y+14), "TOTAL TTC", font=f_large, fill=C_DARK)
    ttc_str = fmt_fcfa(total_ttc)
    try:
        bb = draw.textbbox((0,0), ttc_str, font=f_large)
        tw = bb[2]-bb[0]
    except Exception:
        tw = len(ttc_str)*11
    draw.text((W-tw-18, y+14), ttc_str, font=f_large, fill=C_DARK)
    y += 50

    # Pied de page
    draw.rectangle([(0, H-FOOTER), (W, H)], fill=C_DARK)
    ts  = datetime.now().strftime("%d/%m/%Y %H:%M")
    draw.text((14, H-FOOTER+(FOOTER-11)//2),
              f"Genere le {ts} — GESFI GROUP v4.1", font=f_small, fill=C_ACCENT)
    nom = donnees.get("nom_fichier", "")
    try:
        bb = draw.textbbox((0,0), nom, font=f_small)
        nw = bb[2]-bb[0]
    except Exception:
        nw = len(nom)*6
    draw.text((W-nw-14, H-FOOTER+(FOOTER-11)//2), nom, font=f_small, fill=C_WHITE)

    buf = io.BytesIO()
    img.save(buf, format="PNG", dpi=(150, 150))
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")

# ── Excel ─────────────────────────────────────────────────────────────────────

def generer_excel_batch(tous_resultats: list) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for resultat in tous_resultats:
        if not resultat.get("succes"):
            continue
        donnees = resultat["donnees"]
        ws = wb.create_sheet(title=Path(resultat["nom"]).stem[:31])

        for col, h in enumerate(["LIBELLE", "VALEUR"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", start_color="1B3A5C")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 52

        rows = [
            ("TYPE FACTURE",    donnees.get("type","")),
            ("SOLDE",           "OUI" if donnees.get("solde") else "NON"),
            ("N FACTURE",       donnees.get("num_facture","")),
        ]
        if donnees.get("type") == "VENTE":
            rows += [
                ("Date emission",    donnees.get("date_emission","")),
                ("Date echeance",    donnees.get("date_echeance","")),
                ("Source",           donnees.get("source","")),
                ("RCCM",             donnees.get("rccm","")),
                ("Capital",          donnees.get("capital","")),
                ("Client",           donnees.get("client","")),
                ("Mode de paiement", donnees.get("paiement","")),
            ]
        else:
            rows += [
                ("Date",                donnees.get("date","")),
                ("Livraison prevue",    donnees.get("livraison","")),
                ("Source",              donnees.get("source","")),
                ("Fournisseur Nom",     donnees.get("fournisseur_nom","")),
                ("Fournisseur Adresse", donnees.get("fournisseur_adresse","")),
                ("Fournisseur Tel",     donnees.get("fournisseur_tel","")),
                ("Fournisseur Email",   donnees.get("fournisseur_email","")),
            ]
        rows += [
            ("Description",              donnees.get("description","")),
            ("Quantite",                 donnees.get("quantite_str","")),
            ("Prix unitaire HT (FCFA)",  donnees.get("prix_ht", 0)),
            ("Sous-total HT (FCFA)",     donnees.get("sous_total_ht", 0)),
            (f"TVA ({donnees.get('tva_pct',0):.0f}%) (FCFA)", donnees.get("montant_tva", 0)),
            ("Total TTC (FCFA)",         donnees.get("total_ttc", 0)),
        ]

        r_idx = 2
        for libelle, valeur in rows:
            if valeur == "" or valeur is None:
                continue
            ws.cell(row=r_idx, column=1, value=libelle).font = Font(bold=True)
            ws.cell(row=r_idx, column=2, value=valeur)
            for c in [1, 2]:
                ws.cell(row=r_idx, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            r_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Traitement ────────────────────────────────────────────────────────────────

def traiter_un_fichier(fichier_info: tuple) -> dict:
    idx, fichier, solde = fichier_info
    try:
        if not extension_autorisee(fichier.filename):
            return {"index": idx, "nom": fichier.filename,
                    "erreur": "Format non supporte. Seuls PNG et PDF sont acceptes."}
        img_bytes = fichier.read()
        fichier.seek(0)
        texte = extraire_texte(img_bytes, fichier.filename)
        if not texte.strip():
            return {"index": idx, "nom": fichier.filename, "erreur": "Aucun texte detecte"}
        donnees   = parser_facture_complet(texte, fichier.filename, solde=solde)
        image_b64 = generer_image_facture(donnees)
        return {"index": idx, "nom": fichier.filename, "succes": True,
                "donnees": donnees, "image_b64": image_b64}
    except Exception as e:
        return {"index": idx, "nom": fichier.filename, "erreur": str(e)}

# ── Routes Flask ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", tesseract_ok=TESSERACT_OK, couleurs=COULEURS_GESFI)

@app.route("/analyser_batch", methods=["POST"])
def analyser_batch():
    if not TESSERACT_OK:
        return jsonify({"erreur": "Tesseract non installe"}), 500
    if "fichiers" not in request.files:
        return jsonify({"erreur": "Aucun fichier recu"}), 400
    fichiers = request.files.getlist("fichiers")
    soldes = [request.form.get(f"solde_{i}", "false").lower() == "true"
              for i in range(len(fichiers))]
    if len(fichiers) > 100:
        return jsonify({"erreur": "Maximum 100 fichiers autorises"}), 400
    if len(fichiers) == 0:
        return jsonify({"erreur": "Aucun fichier selectionne"}), 400

    fichiers_a_traiter = [
        (i, f, soldes[i] if i < len(soldes) else False)
        for i, f in enumerate(fichiers) if f.filename
    ]
    resultats = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(traiter_un_fichier, ft): ft for ft in fichiers_a_traiter}
        for future in as_completed(futures):
            resultats.append(future.result())
    resultats.sort(key=lambda x: x.get("index", 0))

    return jsonify({
        "total":     len(resultats),
        "succes":    len([r for r in resultats if r.get("succes")]),
        "erreurs":   len([r for r in resultats if r.get("erreur")]),
        "resultats": resultats,
    })

@app.route("/telecharger_excel", methods=["POST"])
def telecharger_excel():
    data = request.get_json()
    if not data or "resultats" not in data:
        return jsonify({"erreur": "Donnees manquantes"}), 400
    excel_bytes = generer_excel_batch(data.get("resultats", []))
    return send_file(
        io.BytesIO(excel_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"factures_extraites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )

@app.route("/sante")
def sante():
    return jsonify({
        "status": "ok", "tesseract": TESSERACT_OK, "version": "4.1.0",
        "formats_supportes": ["PNG", "PDF"],
        "timestamp": datetime.now().isoformat(),
    })

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
